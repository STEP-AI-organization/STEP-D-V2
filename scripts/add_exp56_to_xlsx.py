"""통합본 xlsx에 Exp 5·6 시트 추가.
- 06_시각훅_194: exp5_haha_visual_features.csv 그대로
- 07_시각훅_프로파일: exp5 profile.json 요약(tier 대조표)
- 08_리텐션조인_71: exp6_retention_join.json (구간 유지율)
- 09_리텐션IoU_71: exp6_retention_iou.json (상위 20% ∩ 채택 IoU)
- 05_파일명세: 신규 4행 append
- 개요·통계: 신규 실험 요약 append
"""
import csv
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding="utf-8")

REPO = Path(r"C:\Users\STEPAI05\STEPD-repo")
XLSX = REPO / "바우처_결과보고_2026" / "제출서류_완성" / "[스텝에이아이] 결과물 증빙 데이터셋_통합.xlsx"
DATA = REPO / "바우처_결과보고_2026" / "증빙_데이터셋" / "실험자료"

wb = load_workbook(XLSX)

# 기존 신규 시트 있으면 제거 (재실행 대비)
for name in ("06_시각훅_194", "07_시각훅_프로파일", "08_리텐션조인_71", "09_리텐션IoU_71"):
    if name in wb.sheetnames:
        del wb[name]

thin = Side(style="thin", color="C0C0C0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="4472C4")


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


# --- 06_시각훅_194 (Exp 5 raw features) ---
ws = wb.create_sheet("06_시각훅_194")
with open(DATA / "exp5_haha_visual_features.csv", encoding="utf-8-sig") as f:
    reader = csv.reader(f)
    for i, row in enumerate(reader, 1):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
style_header(ws, len(row))
# 열 폭
widths = {1: 14, 2: 12, 3: 8, 4: 10, 5: 12, 6: 40, 7: 15, 8: 15, 9: 8, 10: 15, 11: 15, 12: 12, 13: 60}
for col, w in widths.items():
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
ws.freeze_panes = "A2"

# --- 07_시각훅_프로파일 (Exp 5 tier 대조 요약) ---
ws = wb.create_sheet("07_시각훅_프로파일")
profile = json.load(open(DATA / "exp5_haha_visual_profile.json", encoding="utf-8"))

r = 1
ws.cell(row=r, column=1, value="[Exp 5] 하하PD 숏폼 194편 첫 3초 시각 요소 학습 — tier 대조 요약").font = Font(bold=True, size=13)
r += 1
ws.cell(row=r, column=1, value=f"표본: high {profile['sample']['high']} · mid {profile['sample']['mid']} · low {profile['sample']['low']} · OCR 완료 {profile['sample']['ocr_covered']}").font = Font(italic=True, color="595959")
r += 2

# 수치 특성 (tier별 평균)
ws.cell(row=r, column=1, value="■ 수치 특성 (tier별 평균, lift = high/low)").font = Font(bold=True, color="4472C4")
r += 1
headers = ["feature", "high", "mid", "low", "lift(h/l)", "판정"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")
r += 1
for feat, vals in profile["num_features"].items():
    lift = vals.get("lift", 0)
    judg = "🥇 최강" if lift >= 1.5 else "🥈 강" if lift >= 1.2 else "약" if 0.85 <= lift <= 1.15 else "역신호" if lift < 0.85 else "-"
    ws.cell(row=r, column=1, value=feat)
    ws.cell(row=r, column=2, value=round(vals["high"], 3))
    ws.cell(row=r, column=3, value=round(vals["mid"], 3))
    ws.cell(row=r, column=4, value=round(vals["low"], 3))
    ws.cell(row=r, column=5, value=lift)
    ws.cell(row=r, column=6, value=judg)
    r += 1
r += 1

# 훅 타입
ws.cell(row=r, column=1, value="■ 시각 훅 타입 (high-low diff, %p)").font = Font(bold=True, color="4472C4")
r += 1
for i, h in enumerate(["hook_type", "high(%)", "low(%)", "diff(%p)", "판정"], 1):
    c = ws.cell(row=r, column=i, value=h); c.font = header_font; c.fill = header_fill
r += 1
for hook, vals in profile["visual_hook_signals"].items():
    diff = round(vals["diff"] * 100, 1)
    judg = "🥇 강" if diff >= 10 else "🥈 약간" if diff >= 5 else "회피 신호" if diff <= -5 else "무신호"
    ws.cell(row=r, column=1, value=hook)
    ws.cell(row=r, column=2, value=round(vals["high"] * 100, 1))
    ws.cell(row=r, column=3, value=round(vals["low"] * 100, 1))
    ws.cell(row=r, column=4, value=diff)
    ws.cell(row=r, column=5, value=judg)
    r += 1
r += 1

# 색상
ws.cell(row=r, column=1, value="■ 지배 색상 (high-low diff, %p)").font = Font(bold=True, color="4472C4")
r += 1
for i, h in enumerate(["color", "high(%)", "low(%)", "diff(%p)", "판정"], 1):
    c = ws.cell(row=r, column=i, value=h); c.font = header_font; c.fill = header_fill
r += 1
for color, vals in profile["dominant_color_signals"].items():
    diff = round(vals["diff"] * 100, 1)
    judg = "🥇 강" if diff >= 5 else "약" if diff >= 2 else "회피 신호" if diff <= -5 else "무신호"
    ws.cell(row=r, column=1, value=color)
    ws.cell(row=r, column=2, value=round(vals["high"] * 100, 1))
    ws.cell(row=r, column=3, value=round(vals["low"] * 100, 1))
    ws.cell(row=r, column=4, value=diff)
    ws.cell(row=r, column=5, value=judg)
    r += 1
r += 2

# 학습 반영 힌트
ws.cell(row=r, column=1, value="■ 학습 반영 힌트 (recommend에 주입할 신호)").font = Font(bold=True, color="4472C4")
r += 1
hints = profile["recommend_hints"]
for label, val in [
    ("선호 훅 타입", ", ".join(hints["prefer_hook_types"])),
    ("회피 훅 타입", ", ".join(hints["avoid_hook_types"])),
    ("선호 색상", ", ".join(hints["prefer_colors"])),
    ("회피 색상", ", ".join(hints["avoid_colors"]) or "(없음)"),
    ("얼굴 클로즈업 선호?", "네" if hints["prefer_face_close"] else "아니오"),
    ("오버레이 텍스트 선호?", "네" if hints["prefer_overlay"] else "아니오"),
]:
    ws.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws.cell(row=r, column=2, value=val)
    r += 1

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 14

# --- 08_리텐션조인_71 (Exp 6 A-2) ---
ws = wb.create_sheet("08_리텐션조인_71")
join_data = json.load(open(DATA / "exp6_retention_join.json", encoding="utf-8"))
cols = ["short_videoId", "long_videoId", "short_title", "views", "tier",
        "seg_start_sec", "seg_end_sec", "seg_len_sec", "long_dur_min", "n_points",
        "seg_avg_watch", "seg_min_watch", "seg_max_watch", "whole_avg_watch", "rel_vs_whole"]
for i, h in enumerate(cols, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, len(cols))
for r_idx, row in enumerate(join_data, 2):
    ws.cell(row=r_idx, column=1, value=row["short"])
    ws.cell(row=r_idx, column=2, value=row["long"])
    ws.cell(row=r_idx, column=3, value=row["title"])
    ws.cell(row=r_idx, column=4, value=row["views"])
    ws.cell(row=r_idx, column=5, value=row["tier"])
    ws.cell(row=r_idx, column=6, value=row["seg_start"])
    ws.cell(row=r_idx, column=7, value=row["seg_end"])
    ws.cell(row=r_idx, column=8, value=row["seg_len"])
    ws.cell(row=r_idx, column=9, value=row["long_dur_min"])
    ws.cell(row=r_idx, column=10, value=row["n_points"])
    ws.cell(row=r_idx, column=11, value=float(row["seg_avg_watch"]))
    ws.cell(row=r_idx, column=12, value=float(row["seg_min_watch"]))
    ws.cell(row=r_idx, column=13, value=float(row["seg_max_watch"]))
    ws.cell(row=r_idx, column=14, value=float(row["whole_avg_watch"]))
    ws.cell(row=r_idx, column=15, value=float(row["rel_vs_whole"]))

for col, w in zip("ABCDEFGHIJKLMNO", [14, 14, 50, 10, 8, 10, 10, 10, 12, 10, 14, 14, 14, 15, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- 09_리텐션IoU_71 (Exp 6 A-4) ---
ws = wb.create_sheet("09_리텐션IoU_71")
iou_data = json.load(open(DATA / "exp6_retention_iou.json", encoding="utf-8"))
cols = ["short_videoId", "long_videoId", "tier", "seg_start_sec", "seg_end_sec",
        "bestIoU (with top-20% retention)", "anyOverlap", "topSpansCount"]
for i, h in enumerate(cols, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, len(cols))
for r_idx, row in enumerate(iou_data, 2):
    ws.cell(row=r_idx, column=1, value=row["short"])
    ws.cell(row=r_idx, column=2, value=row["long"])
    ws.cell(row=r_idx, column=3, value=row["tier"])
    ws.cell(row=r_idx, column=4, value=row["seg"][0])
    ws.cell(row=r_idx, column=5, value=row["seg"][1])
    ws.cell(row=r_idx, column=6, value=row["bestIoU"])
    ws.cell(row=r_idx, column=7, value="Y" if row["anyOverlap"] else "N")
    ws.cell(row=r_idx, column=8, value=row["topSpansCount"])
for col, w in zip("ABCDEFGH", [14, 14, 8, 12, 12, 30, 12, 14]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- 05_파일명세 append ---
ws = wb["05_파일명세"]
last = ws.max_row
new_entries = [
    ("exp5_haha_visual_features.csv", "06_시각훅_194", 194, "하하PD 숏폼 194편 첫 3초 시각특성 (Gemini Vision 8특성 판정 · Exp 5)"),
    ("exp5_haha_visual_profile.json", "07_시각훅_프로파일", 1, "Exp 5 tier 대조 요약 프로파일 (수치·훅·색상 diff + recommend 힌트)"),
    ("exp6_retention_join.json", "08_리텐션조인_71", 71, "매칭 71쌍 × 롱폼 유튜브 리텐션 커브 조인 — 채택 구간 유지율 산출 (Exp 6 A-2·3)"),
    ("exp6_retention_iou.json", "09_리텐션IoU_71", 71, "리텐션 상위 20% 구간 vs 편집자 채택 구간 IoU 판정 (Exp 6 A-4)"),
]
for i, (fname, sheet, n, desc) in enumerate(new_entries, 1):
    r = last + i
    ws.cell(row=r, column=1, value=fname)
    ws.cell(row=r, column=2, value=sheet)
    ws.cell(row=r, column=3, value=n)
    ws.cell(row=r, column=4, value=desc)

# --- 개요·통계 하단에 Exp 5·6 요약 append ---
ws = wb["개요·통계"]
last = ws.max_row
r = last + 2
ws.cell(row=r, column=1, value="■ Exp 5·6 추가 실험 요약 (2026-07-22)").font = Font(bold=True, size=12, color="4472C4")
r += 1
for line in [
    "Exp 5 (시각훅 194편): text_cue +13%p · reaction +10%p · situation -7%p (회피) · n_faces lift 1.58 · 흰색 lift 4.0",
    "  → \"첫 3초에 텍스트 큐 + 얼굴 2명+ + 밝은/흰색 + 리액션\"이 오프닝 훅 공식 · 잔잔한 상황설정은 회피 신호",
    "Exp 6 (리텐션 71쌍): high 구간 유지율 1.02배 · low 0.88배 · IoU 0.068 (목표 0.3 미달)",
    "  → 편집자가 잘 자른 구간은 시청자도 실제 잘 봤음 (감 검증). 리텐션은 픽 신호로는 약함, 회피 필터로만 유효",
]:
    ws.cell(row=r, column=1, value=line)
    r += 1

wb.save(XLSX)
print("완료. 신규 시트 4개 + 05_파일명세 4행 + 개요·통계 요약 append")
print("시트 목록:", wb.sheetnames)
