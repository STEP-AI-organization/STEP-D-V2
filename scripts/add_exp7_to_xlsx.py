"""통합본 xlsx에 Exp 7 결과 시트 추가."""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding="utf-8")
REPO = Path(r"C:\Users\STEPAI05\STEPD-repo")
XLSX = REPO / "바우처_결과보고_2026" / "제출서류_완성" / "[스텝에이아이] 결과물 증빙 데이터셋_통합.xlsx"
DATA = REPO / "바우처_결과보고_2026" / "증빙_데이터셋" / "실험자료"

wb = load_workbook(XLSX)
if "10_시각훅AB_Exp7" in wb.sheetnames:
    del wb["10_시각훅AB_Exp7"]

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="4472C4")
warn_fill = PatternFill("solid", fgColor="FFF2CC")

ws = wb.create_sheet("10_시각훅AB_Exp7")
r = 1
ws.cell(row=r, column=1, value="[Exp 7] 시각훅 프로파일 recommend 반영 A/B (Phase 5d, 2026-07-22)").font = Font(bold=True, size=13)
r += 1
ws.cell(row=r, column=1, value="Exp 5(194편 대조)로 발견한 시각훅 공식을 recommend 프롬프트에 실제 주입해 Topic-Hit@N 개선 여부 실증").font = Font(italic=True, color="595959")
r += 1
ws.cell(row=r, column=1, value="표본: 3홀드아웃(원정대5 · 원정대4 · 경주PC방) · 정답 26편 · runs 1회 (통계 미확정)").font = Font(italic=True, color="595959")
r += 2

# 결과표
ws.cell(row=r, column=1, value="■ Topic-Hit@N (내용 일치, Gemini 판정)").font = Font(bold=True, color="4472C4")
r += 1
for i, h in enumerate(["조건", "Topic@5", "Topic@10", "Topic@20", "판정 (@10)"], 1):
    c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
r += 1

result = json.load(open(DATA / "exp7_visual_ab.json", encoding="utf-8"))
rows = [
    ("off (학습 없음)", result["profile_off"], "기준"),
    ("base (기존 학습 규칙)", result["profile_a"], "base = off (변화 없음)"),
    ("base+시각훅 (Exp 5 반영)", result["profile_b"], "⚠️ -3.8%p 저하"),
]
for label, d, judg in rows:
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=round(d["5"]["mean"], 3))
    ws.cell(row=r, column=3, value=round(d["10"]["mean"], 3))
    ws.cell(row=r, column=4, value=round(d["20"]["mean"], 3))
    ws.cell(row=r, column=5, value=judg)
    if "시각훅" in label:
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = warn_fill
    r += 1

r += 1
ws.cell(row=r, column=1, value="■ 핵심 발견 (예상 뒤집힘)").font = Font(bold=True, color="C00000")
r += 1
for line in [
    "1. 시각훅 프로파일 프롬프트 주입 → 오히려 성능 저하 (@10 base 0.500 → base+visual 0.462, -3.8%p)",
    "2. base+visual = off보다도 열등 (Exp 3의 '학습 켜면 Topic에 해' 패턴이 시각훅 추가로 강화)",
    "3. --runs 1이라 σ=0 · Exp 3에서 σ~0.06~0.08 관측된 걸 감안하면 통계 확정은 5회 반복 필요",
]:
    ws.cell(row=r, column=1, value=line); r += 1
r += 1
ws.cell(row=r, column=1, value="■ 원인 가설").font = Font(bold=True, color="4472C4")
r += 1
for line in [
    "(a) 프롬프트 과부하 — base 이미 1552자, 시각훅 5줄 추가로 지시 상충",
    "(b) 다양성 저해 — 'reaction·text_cue 우대'가 다른 좋은 순간을 배제 (Exp 3 hookWeights 편중 패턴)",
    "(c) 신호 형식 불일치 — Exp 5는 완성된 숏폼 첫 3초 통계 · 롱폼 장면 텍스트로 '오프닝 훅' 판단 어려움",
]:
    ws.cell(row=r, column=1, value=line); r += 1
r += 1
ws.cell(row=r, column=1, value="■ 다음 액션").font = Font(bold=True, color="4472C4")
r += 1
for line in [
    "1. 5회 반복 재실행 → σ 확정 · 통계 확정",
    "2. 학습 프로파일 재설계 — hookWeights를 후보 생성에서 제거하고 랭킹 소량 부스팅으로만 이동 (트랙 2, 계획서 §3-1)",
    "3. 시각훅 신호 다른 형태로 시도: (i) 랭킹 페널티(situation 감점)만, (ii) vision 단계 태깅 후 매칭",
]:
    ws.cell(row=r, column=1, value=line); r += 1

for col, w in zip("ABCDE", [30, 12, 14, 12, 30]):
    ws.column_dimensions[col].width = w

# 05_파일명세 append
ws2 = wb["05_파일명세"]
last = ws2.max_row + 1
ws2.cell(row=last, column=1, value="exp7_visual_ab.json")
ws2.cell(row=last, column=2, value="10_시각훅AB_Exp7")
ws2.cell(row=last, column=3, value=1)
ws2.cell(row=last, column=4, value="Exp 7 시각훅 A/B — 3조건(off/base/base+visual) × 3홀드 × 1회 Topic-Hit@N")

# 개요·통계 하단에 append
ws3 = wb["개요·통계"]
last = ws3.max_row + 2
ws3.cell(row=last, column=1, value="■ Exp 7 (시각훅 recommend 반영) — 예상 뒤집힘").font = Font(bold=True, size=12, color="C00000")
last += 1
for line in [
    "Topic@10: off 0.500 · base 0.500 · base+visual 0.462 → 시각훅 프롬프트 주입이 오히려 성능 저하",
    "  → Exp 5(패턴 실증) → Exp 7(주입 실패) · 다음: 5회 반복 확정 + hookWeights 재설계(트랙 2)",
]:
    ws3.cell(row=last, column=1, value=line); last += 1

wb.save(XLSX)
print("완료. 시트 10_시각훅AB_Exp7 추가")
print("시트 목록:", wb.sheetnames)
