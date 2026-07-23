"""통합 xlsx에 댓글 원본·추출 시트 추가 (데이터 보존 가치)."""
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8")
REPO = Path(r"C:\Users\STEPAI05\STEPD-repo")
XLSX = REPO / "바우처_결과보고_2026" / "제출서류_완성" / "[스텝에이아이] 결과물 증빙 데이터셋_통합.xlsx"
DATA = REPO / "바우처_결과보고_2026" / "증빙_데이터셋" / "실험자료"

wb = load_workbook(XLSX)

sheets_to_add = [
    ("19_댓글원본_Exp10", "댓글원본_Exp10_15롱폼.csv"),
    ("20_댓글추출_Exp10", "댓글Gemini추출_Exp10_15롱폼.csv"),
    ("21_댓글원본_Exp11", "댓글원본_Exp11_하하과거8편.csv"),
    ("22_댓글추출_Exp11", "댓글Gemini추출_Exp11_하하과거8편.csv"),
    ("23_시청자지목시간", "시청자지목시간_Exp10B안.csv"),
    ("24_v2winner_시청자매칭", "v2winner_시청자매칭_Exp10.5.csv"),
]

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="4472C4")
star_fill = PatternFill("solid", fgColor="FFEB9C")

for sheet_name, csv_name in sheets_to_add:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    csv_path = DATA / csv_name
    if not csv_path.exists():
        print(f"  · {csv_name} 없음 skip")
        continue
    rows = list(csv.DictReader(open(csv_path, encoding="utf-8-sig")))
    if not rows:
        continue
    ws = wb.create_sheet(sheet_name)
    fields = list(rows[0].keys())
    for i, h in enumerate(fields, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, k in enumerate(fields, 1):
            v = row.get(k, "")
            # 숫자 변환 시도
            if k in ("likes", "sec", "comment_idx", "winner_idx", "hint_likes"):
                try:
                    v = int(v) if v not in ("", None) else v
                except (ValueError, TypeError):
                    pass
            elif k in ("moment_ref", "quote_ref", "demand", "matched", "is_pinned"):
                if v == "True": v = True
                elif v == "False": v = False
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            # 좋아요 100+ 강조
            if k == "likes" and isinstance(v, int) and v >= 100:
                c.font = Font(bold=True, color="C00000")
                c.fill = star_fill
    ws.freeze_panes = "A2"
    # 컬럼 너비 자동
    for i, k in enumerate(fields, 1):
        col = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        if k == "text" or k == "context_text" or k == "reasoning":
            ws.column_dimensions[col].width = 60
        elif k == "long_title" or k == "title" or k == "best_viewer_hint" or k == "moment_hint" or k == "demand_text":
            ws.column_dimensions[col].width = 35
        elif k in ("longVideoId", "raw_time_notation"):
            ws.column_dimensions[col].width = 15
        elif k in ("channel", "emotion", "sentiment", "moment_type", "demand_category", "hook", "mmss"):
            ws.column_dimensions[col].width = 12
        else:
            ws.column_dimensions[col].width = 10
    print(f"  ✓ {sheet_name}: {len(rows)} rows")

# 07_파일명세 append
ws2 = wb["07_파일명세"]
last = ws2.max_row + 1
for fname, sheet, n, desc in [
    ("댓글원본_Exp10_15롱폼.csv", "19_댓글원본_Exp10", 358, "3채널 15롱폼 상위 좋아요 원본 댓글"),
    ("댓글Gemini추출_Exp10_15롱폼.csv", "20_댓글추출_Exp10", 358, "위 원본에 Gemini 8필드 분류"),
    ("댓글원본_Exp11_하하과거8편.csv", "21_댓글원본_Exp11", 350, "하하 과거 7 롱폼 학습용 원본 댓글"),
    ("댓글Gemini추출_Exp11_하하과거8편.csv", "22_댓글추출_Exp11", 350, "Exp 11 Gemini 세분화 카테고리 (moment_type · demand_category)"),
    ("시청자지목시간_Exp10B안.csv", "23_시청자지목시간", 12, "시청자 명시 M:SS 시간 표기 · 좋아요·맥락"),
    ("v2winner_시청자매칭_Exp10.5.csv", "24_v2winner_시청자매칭", 26, "v2 winners 26개 vs 시청자 상위 목소리 매칭 판정"),
]:
    ws2.cell(row=last, column=1, value=fname)
    ws2.cell(row=last, column=2, value=sheet)
    ws2.cell(row=last, column=3, value=n)
    ws2.cell(row=last, column=4, value=desc)
    last += 1

# 개요·통계 append
ws3 = wb["개요·통계"]
last = ws3.max_row + 2
ws3.cell(row=last, column=1, value="■ 댓글 데이터 CSV 아카이브 (Exp 10·11 데이터 보존)").font = Font(bold=True, size=12, color="4472C4")
last += 1
for line in [
    "6개 CSV · 총 1,454 rows (원본 댓글 · Gemini 추출 · 시청자 지목 시간 · winner 매칭 판정)",
    "  Exp 10 15롱폼 358행 · Exp 11 하하 7편 350행 · timestamps 12행 · winner 매칭 26행",
    "  향후 신호 튜닝·모델 학습·B2B 확장 base 데이터셋 · NCC 보고서 실증 근거",
]:
    ws3.cell(row=last, column=1, value=line); last += 1

wb.save(XLSX)
print(f"\n완료. 총 {len(wb.sheetnames)} 시트")
print("전체 시트:", wb.sheetnames)
