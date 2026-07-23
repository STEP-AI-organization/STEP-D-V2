"""통합 xlsx에 Exp 10 시청자 신호 시트 추가."""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8")
REPO = Path(r"C:\Users\STEPAI05\STEPD-repo")
XLSX = REPO / "바우처_결과보고_2026" / "제출서류_완성" / "[스텝에이아이] 결과물 증빙 데이터셋_통합.xlsx"
RES = Path("D:/STEPD-experiments/results")
REPORT_DIR = REPO / "바우처_결과보고_2026" / "증빙_데이터셋" / "실험자료"

# 사본
import shutil
for name in ("exp10_all_comments.json", "exp10_all_extracted.json", "exp10_all_analysis.json"):
    src = RES / name
    if src.exists():
        shutil.copy2(src, REPORT_DIR / name)

wb = load_workbook(XLSX)
for name in ("13_시청자신호_15롱폼", "14_시청자상위목소리"):
    if name in wb.sheetnames:
        del wb[name]

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="4472C4")
warn_fill = PatternFill("solid", fgColor="FFEB9C")
ok_fill = PatternFill("solid", fgColor="C6EFCE")
bad_fill = PatternFill("solid", fgColor="FFC7CE")

analysis = json.load(open(RES / "exp10_all_analysis.json", encoding="utf-8"))

# --- 13_시청자신호_15롱폼 ---
ws = wb.create_sheet("13_시청자신호_15롱폼")
r = 1
ws.cell(row=r, column=1, value="[Exp 10] 3채널 15롱폼 시청자 반응·기대 신호 파일럿").font = Font(bold=True, size=14); r += 1
ws.cell(row=r, column=1, value="상위 좋아요 100개 댓글 × 15롱폼 = 1500건 · Gemini 8필드 추출").font = Font(italic=True, color="595959"); r += 2

# 채널 평균표
ws.cell(row=r, column=1, value="■ 채널별 평균 신호 강도").font = Font(bold=True, color="4472C4"); r += 1
hdrs = ["채널", "롱폼 수", "moment_ref", "quote_ref", "demand", "판정"]
for i, h in enumerate(hdrs, 1):
    c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
r += 1

HAHA_LIDS = ["LcMolKaPcrw", "NtXLj7xOeE8", "JppILjNTCok"]
ENA_LIDS = ["dnIaj6L3t1E", "DPclbGO1F9g", "Lj_tFgRqqEI", "MjWwq8bBwJE", "QNtoQ4zI8mc"]
DNA_LIDS = ["rhX9po-DBZI", "NUM1zfQujWY", "OuvpspSaAUQ", "k8BHuiKF0rk", "ALuFb_TqHPU", "a9O8d0zLfTg", "sT9KQTLg2Cs"]

for ch_name, lids, fill, judge in [
    ("하하", HAHA_LIDS, ok_fill, "⭐ 강한 신호"),
    ("ENA", ENA_LIDS, warn_fill, "🔵 유효 신호"),
    ("드나드나", DNA_LIDS, bad_fill, "❌ 신호 부족"),
]:
    entries = [analysis.get(lid) for lid in lids if analysis.get(lid)]
    if not entries:
        continue
    n = len(entries)
    m = sum(e.get("moment_ref_pct", 0) for e in entries) / n
    q = sum(e.get("quote_ref_pct", 0) for e in entries) / n
    d = sum(e.get("demand_pct", 0) for e in entries) / n
    ws.cell(row=r, column=1, value=ch_name).fill = fill
    ws.cell(row=r, column=2, value=n).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3, value=f"{m:.1f}%").alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value=f"{q:.1f}%").alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=5, value=f"{d:.1f}%").alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=6, value=judge).alignment = Alignment(horizontal="center")
    r += 1
r += 1

# 롱폼별 세부
ws.cell(row=r, column=1, value="■ 롱폼별 세부").font = Font(bold=True, color="4472C4"); r += 1
hdrs = ["채널", "롱폼", "댓글수", "moment_ref%", "quote_ref%", "demand%"]
for i, h in enumerate(hdrs, 1):
    c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
r += 1
for ch_name, lids in [("하하", HAHA_LIDS), ("ENA", ENA_LIDS), ("드나드나", DNA_LIDS)]:
    for lid in lids:
        e = analysis.get(lid)
        if not e:
            ws.cell(row=r, column=1, value=ch_name).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=2, value=lid)
            ws.cell(row=r, column=3, value="—").alignment = Alignment(horizontal="center")
            for c_i in [4, 5, 6]:
                ws.cell(row=r, column=c_i, value="—").alignment = Alignment(horizontal="center")
            r += 1
            continue
        ws.cell(row=r, column=1, value=ch_name).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=lid)
        ws.cell(row=r, column=3, value=e.get("n", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=f"{e.get('moment_ref_pct',0):.1f}%").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5, value=f"{e.get('quote_ref_pct',0):.1f}%").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6, value=f"{e.get('demand_pct',0):.1f}%").alignment = Alignment(horizontal="center")
        r += 1
r += 1

# 결론
ws.cell(row=r, column=1, value="■ 결론").font = Font(bold=True, color="4472C4"); r += 1
for line in [
    "1. **채널 규모 = 시청자 신호 강도** — 인기 채널(하하 56%) → 유효(ENA 16%) → 부족(드나드나 14%, 댓글 0~1개 대다수)",
    "2. **파이프라인 반영 시 채널별 임계 필요** — 예: 롱폼당 댓글 30+ 채널에만 viewer_moment_density 신호 적용",
    "3. **명시적 타임스탬프 언급 발견** — ENA '10:43 상황' 등 (start,end) 파싱 가능한 정확 신호 (규모 무관)",
    "4. **v2 winner 시청자 검증** — 하하 원정대5 winner #5 '사장님은 양상국만'과 상위 언급 '양상국 기죽음/웃김/나락' 의미 일치",
    "",
    "**파이프라인 반영 후보 3안**:",
    "  A) viewer_moment_density (v2 6번째 신호, 인기 채널용, 댓글 30+ 임계)",
    "  B) explicit_timestamp 파싱 (규모 무관, 정확한 픽 후보 발굴)",
    "  C) demand 대시보드 (콘텐츠 전략용, 파이프라인 밖)",
]:
    ws.cell(row=r, column=1, value=line); r += 1

for col, w in zip("ABCDEF", [15, 20, 12, 14, 14, 20]):
    ws.column_dimensions[col].width = w

# --- 14_시청자상위목소리 ---
ws = wb.create_sheet("14_시청자상위목소리")
r = 1
ws.cell(row=r, column=1, value="[Exp 10] 채널별 시청자 상위 목소리 (좋아요순)").font = Font(bold=True, size=14); r += 1
ws.cell(row=r, column=1, value="moment (특정 순간 언급) + demand (재출연·후속·요청 등 기대)").font = Font(italic=True, color="595959"); r += 2

for ch_name, lids in [("하하", HAHA_LIDS), ("ENA", ENA_LIDS), ("드나드나", DNA_LIDS)]:
    ws.cell(row=r, column=1, value=f"[{ch_name}]").font = Font(bold=True, size=13, color="4472C4"); r += 1

    # 순간 언급 종합
    all_moments = []
    all_demands = []
    for lid in lids:
        e = analysis.get(lid)
        if not e:
            continue
        all_moments.extend(e.get("top_moments") or [])
        all_demands.extend(e.get("top_demands") or [])
    all_moments.sort(key=lambda t: -(t[1] if isinstance(t, list) and len(t) > 1 else 0))
    all_demands.sort(key=lambda t: -(t[1] if isinstance(t, list) and len(t) > 1 else 0))

    ws.cell(row=r, column=1, value="상위 순간 언급").font = Font(bold=True); r += 1
    hdrs = ["좋아요", "moment_hint"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill
    r += 1
    for m in all_moments[:10]:
        hint = m[0] if isinstance(m, list) else "—"
        likes = m[1] if isinstance(m, list) and len(m) > 1 else 0
        ws.cell(row=r, column=1, value=likes).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=hint)
        if isinstance(likes, int) and likes >= 100:
            ws.cell(row=r, column=1).font = Font(bold=True, color="C00000")
        r += 1

    if all_demands:
        r += 1
        ws.cell(row=r, column=1, value="상위 요청 (Demand)").font = Font(bold=True); r += 1
        for h in hdrs:
            pass
        hdrs2 = ["좋아요", "demand_text"]
        for i, h in enumerate(hdrs2, 1):
            c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill
        r += 1
        for d in all_demands[:8]:
            txt = d[0] if isinstance(d, list) else "—"
            likes = d[1] if isinstance(d, list) and len(d) > 1 else 0
            ws.cell(row=r, column=1, value=likes).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=2, value=txt)
            if isinstance(likes, int) and likes >= 100:
                ws.cell(row=r, column=1).font = Font(bold=True, color="C00000")
            r += 1
    r += 2

for col, w in zip("AB", [12, 60]):
    ws.column_dimensions[col].width = w

# 07_파일명세 append
ws2 = wb["07_파일명세"]
last = ws2.max_row + 1
for fname, sheet, n, desc in [
    ("exp10_all_comments.json", "13_시청자신호_15롱폼", 1500, "Exp 10 · 15 롱폼 × 상위 좋아요 100 댓글"),
    ("exp10_all_extracted.json", "13_시청자신호_15롱폼", 1500, "Gemini 8필드 추출 (moment_ref·emotion·demand 등)"),
    ("exp10_all_analysis.json", "14_시청자상위목소리", 15, "채널별 요약 통계 + 상위 순간·요청 목록"),
]:
    ws2.cell(row=last, column=1, value=fname)
    ws2.cell(row=last, column=2, value=sheet)
    ws2.cell(row=last, column=3, value=n)
    ws2.cell(row=last, column=4, value=desc)
    last += 1

# 개요·통계 append
ws3 = wb["개요·통계"]
last = ws3.max_row + 2
ws3.cell(row=last, column=1, value="■ Exp 10 · 시청자 반응·기대 신호 파일럿 (댓글 → 8필드 추출)").font = Font(bold=True, size=12, color="4472C4")
last += 1
for line in [
    "3채널 15롱폼 · 상위 좋아요 1500 댓글 · Gemini 8필드 분류",
    "  하하 moment_ref 56% · ENA 16% · 드나드나 14% — 채널 규모=시청자 신호 강도 실측",
    "  하하 최상위 요청 [1500❤ ㅈㄸㄸ쑈 진행 여부] · v2 winner 시청자 언급과 의미 일치 실증",
    "  → 파이프라인 개선 방향 전환: 룰 튜닝(약간 개선) → 시청자 목소리 배선(강한 신호 확보)",
]:
    ws3.cell(row=last, column=1, value=line); last += 1

wb.save(XLSX)
print("완료. 시트 13_시청자신호_15롱폼 · 14_시청자상위목소리 추가")
print("전체 시트:", wb.sheetnames)
