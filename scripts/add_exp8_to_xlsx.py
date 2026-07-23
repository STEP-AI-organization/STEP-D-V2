"""통합본 xlsx에 Exp 8 v2 결정론 결과 시트 추가."""
import csv
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8")
REPO = Path(r"C:\Users\STEPAI05\STEPD-repo")
XLSX = REPO / "바우처_결과보고_2026" / "제출서류_완성" / "[스텝에이아이] 결과물 증빙 데이터셋_통합.xlsx"
DATA = REPO / "바우처_결과보고_2026" / "증빙_데이터셋" / "실험자료"

wb = load_workbook(XLSX)
for name in ("11_히든젬_46픽", "12_히든젬_확정8"):
    if name in wb.sheetnames:
        del wb[name]

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="4472C4")
pass_fill = PatternFill("solid", fgColor="C6EFCE")
warn_fill = PatternFill("solid", fgColor="FFC7CE")

# --- 11_히든젬_46픽 (전체 CSV) ---
ws = wb.create_sheet("11_히든젬_46픽")
rows = list(csv.DictReader(open(DATA / "exp8_v2_deterministic_gems.csv", encoding="utf-8-sig")))
if rows:
    fields = list(rows[0].keys())
    for i, h in enumerate(fields, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, k in enumerate(fields, 1):
            v = row[k]
            if k.startswith("pass_") or k == "all_pass":
                v = (v == "True")
                c = ws.cell(row=r_idx, column=c_idx, value="✅" if v else "❌")
                c.alignment = Alignment(horizontal="center")
                if v:
                    c.fill = pass_fill
            else:
                # 숫자 변환 시도
                try:
                    v = float(v) if "." in str(v) or k in ("appeal",) else int(v) if str(v).lstrip("-").isdigit() else v
                except Exception:
                    pass
                ws.cell(row=r_idx, column=c_idx, value=v)
        if row.get("all_pass") == "True":
            for c_idx in range(1, len(fields)+1):
                if not ws.cell(row=r_idx, column=c_idx).fill.fgColor.rgb or ws.cell(row=r_idx, column=c_idx).fill.fgColor.rgb == "00000000":
                    ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJKLMNOPQRST", [14, 10, 10, 8, 40, 10, 8, 45, 10, 12, 12, 10, 10, 10, 10, 10, 10, 10, 10, 8]):
        ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- 12_히든젬_확정8 (요약 시트) ---
ws = wb.create_sheet("12_히든젬_확정8")
r = 1
ws.cell(row=r, column=1, value="[Exp 8 v2] 편집자 미발견 히든젬 확정 (LLM 판정 X · 결정론적 5신호 통과)").font = Font(bold=True, size=13)
r += 1
ws.cell(row=r, column=1, value="사용자 통찰: '편집자 자른 게 정답은 절대 아니다' — Exp 6에서 편집자 27%(low tier) 오답 실측. STEP D 진짜 가치는 '편집자 놓친 것 잡기'").font = Font(italic=True, color="595959")
r += 2

ws.cell(row=r, column=1, value="■ 5신호 (모두 실측/산술, LLM 판정 X)").font = Font(bold=True, color="4472C4")
r += 1
for line in [
    "1. IoU ≤ 0.1 (편집자 미발견)                     - 산술",
    "2. 리텐션 rel_vs_whole ≥ 1.05                    - 실측 (유튜브 애널리틱스)",
    "3. 길이 30~60초 (플랫폼 최적)                    - 산술",
    "4. 텍스트 밀도 ≥ 3.0 chars/s (침묵 아님·대사)   - 실측 (STT)",
    "5. 훅 ∈ {돌직구·반전·감정고조·갈등·웃음·공감}   - 태그 규칙 (Exp 2 실증)",
]:
    ws.cell(row=r, column=1, value=line); r += 1
r += 1

ws.cell(row=r, column=1, value=f"■ 확정 8개 (5신호 AND 통과 · 46 픽 → 8개 = 17%)").font = Font(bold=True, color="4472C4")
r += 1
headers = ["#", "롱폼", "구간", "제목", "훅", "리텐션 rel", "밀도(자/s)"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
r += 1
gems = json.load(open(DATA / "exp8_v2_confirmed_gems.json", encoding="utf-8"))
gems.sort(key=lambda g: -g["rel_vs_whole"])
for i, g in enumerate(gems, 1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=g["long"])
    ws.cell(row=r, column=3, value=f"{g['start']:.0f}s~{g['end']:.0f}s ({g['seg_len']:.0f}초)")
    ws.cell(row=r, column=4, value=g["title"])
    ws.cell(row=r, column=5, value=g["hook"])
    c = ws.cell(row=r, column=6, value=g["rel_vs_whole"])
    if g["rel_vs_whole"] >= 1.4:
        c.font = Font(bold=True, color="C00000")
    ws.cell(row=r, column=7, value=g["text_density_chars_per_s"])
    r += 1

r += 1
ws.cell(row=r, column=1, value="■ 결론").font = Font(bold=True, color="4472C4")
r += 1
for line in [
    "1. LLM 판정 없이 재현 가능한 8개 히든젬 확정 (근거 5개 명확)",
    "2. v1 Gemini 판정 9개와 6개 겹침 → 근거 다른 방법이 대체로 동일 결론 → 실측 신호 신뢰성 방증",
    "3. 최상위 '경주 마스터의 꽐라 진실게임 개망신 썰' rel 1.75배 — 편집자 아예 안 자른 구간, 시청자는 롱폼 평균보다 75% 더 오래 봄",
    "4. STEP D의 판매 논리: '엔진이 편집자 놓친 것을 잡는다' — 실측 근거 확보",
]:
    ws.cell(row=r, column=1, value=line); r += 1

for col, w in zip("ABCDEFG", [5, 14, 20, 45, 10, 14, 12]):
    ws.column_dimensions[col].width = w

# 05_파일명세 append
ws2 = wb["05_파일명세"]
last = ws2.max_row + 1
for fname, sheet, n, desc in [
    ("exp8_v2_deterministic_gems.csv", "11_히든젬_46픽", 46, "Exp 8 v2 결정론 필터 · 46 픽 전체 5신호 부울 + 실측"),
    ("exp8_v2_confirmed_gems.json", "12_히든젬_확정8", 8, "5신호 전부 통과 히든젬 8개 확정 (편집자 미발견·리텐션·길이·밀도·훅)"),
]:
    ws2.cell(row=last, column=1, value=fname)
    ws2.cell(row=last, column=2, value=sheet)
    ws2.cell(row=last, column=3, value=n)
    ws2.cell(row=last, column=4, value=desc)
    last += 1

# 개요·통계 하단 append
ws3 = wb["개요·통계"]
last = ws3.max_row + 2
ws3.cell(row=last, column=1, value="■ Exp 8 v2 · 편집자 미발견 히든젬 확정 (LLM 판정 X)").font = Font(bold=True, size=12, color="4472C4")
last += 1
for line in [
    "46 픽 → 8개 확정 (17%) · 5개 실측/산술 신호 AND 통과 (LLM 판정 완전 제거, 재현 가능)",
    "  최상위: 경주 마스터 롤 눈물 고백 리텐션 rel 1.75배 (편집자 안 자른 구간, 시청자는 75% 더 오래 봄)",
    "  → STEP D 판매 논리 실측: '편집자가 놓친 것을 잡는 AI'",
]:
    ws3.cell(row=last, column=1, value=line); last += 1

wb.save(XLSX)
print("완료. 시트 11_히든젬_46픽 + 12_히든젬_확정8 추가")
print("전체 시트:", wb.sheetnames)
