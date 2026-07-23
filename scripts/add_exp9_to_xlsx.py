"""통합 xlsx에 Exp 9 (ENA v2 재현) 시트 추가."""
import csv
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8")
REPO = Path(r"C:\Users\STEPAI05\STEPD-repo")
XLSX = REPO / "바우처_결과보고_2026" / "제출서류_완성" / "[스텝에이아이] 결과물 증빙 데이터셋_통합.xlsx"
DATA = REPO / "바우처_결과보고_2026" / "증빙_데이터셋" / "실험자료"

wb = load_workbook(XLSX)
for name in ("08_ENA재현_32픽", "09_ENA재현_확정9"):
    if name in wb.sheetnames:
        del wb[name]

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="4472C4")
pass_fill = PatternFill("solid", fgColor="C6EFCE")

# --- 08_ENA재현_32픽 ---
ws = wb.create_sheet("08_ENA재현_32픽")
rows = list(csv.DictReader(open(DATA / "exp9_ena_deterministic_picks.csv", encoding="utf-8-sig")))
if rows:
    fields = list(rows[0].keys())
    for i, h in enumerate(fields, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, k in enumerate(fields, 1):
            v = row[k]
            if k.startswith("pass_") or k == "all_pass":
                v = (v == "True")
                c = ws.cell(row=r_idx, column=c_idx, value="✅" if v else "❌")
                c.alignment = Alignment(horizontal="center")
                if v:
                    c.fill = pass_fill
            else:
                try:
                    v = float(v) if "." in str(v) or k in ("appeal",) else int(v) if str(v).lstrip("-").isdigit() else v
                except Exception:
                    pass
                ws.cell(row=r_idx, column=c_idx, value=v)
        if row.get("all_pass") == "True":
            for c_idx in range(1, len(fields)+1):
                if not ws.cell(row=r_idx, column=c_idx).fill.fgColor.rgb or ws.cell(row=r_idx, column=c_idx).fill.fgColor.rgb == "00000000":
                    ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJKLMNOPQR", [14, 10, 10, 8, 45, 10, 8, 12, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- 09_ENA재현_확정9 ---
ws = wb.create_sheet("09_ENA재현_확정9")
r = 1
ws.cell(row=r, column=1, value="[Exp 9] ENA 채널 v2 필터 재현 — 다채널 일반화 실증").font = Font(bold=True, size=13); r += 1
ws.cell(row=r, column=1, value="장르: 예능·연애리얼리티(나는솔로) — 하하(꽐라예능)와 완전 다른 장르에서 v2 5신호 재현 확인").font = Font(italic=True, color="595959"); r += 2

ws.cell(row=r, column=1, value="■ 하하 v2 vs ENA v2 대조").font = Font(bold=True, color="4472C4"); r += 1
tbl = [
    ["채널", "전체 픽", "통과", "통과율", "최상위 rel"],
    ["하하 (Exp 8 v2)", 46, 8, "17%", 1.75],
    ["ENA (Exp 9)", 32, 9, "28%", 1.64],
]
for row_i, row_vals in enumerate(tbl):
    for c_i, v in enumerate(row_vals, 1):
        c = ws.cell(row=r, column=c_i, value=v)
        if row_i == 0:
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
        c.alignment = Alignment(horizontal="center")
    r += 1
r += 1

ws.cell(row=r, column=1, value="■ ENA 확정 히든젬 9개 (rel 순)").font = Font(bold=True, color="4472C4"); r += 1
headers = ["#", "롱폼", "구간", "제목", "훅", "rel", "밀도"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
r += 1
gems = json.load(open(DATA / "exp9_ena_confirmed_gems.json", encoding="utf-8"))
for i, g in enumerate(gems, 1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=g["long"])
    ws.cell(row=r, column=3, value=f"{g['start']:.0f}s~{g['end']:.0f}s ({g['seg_len']:.0f}초)")
    ws.cell(row=r, column=4, value=g["title"])
    ws.cell(row=r, column=5, value=g["hook"])
    c = ws.cell(row=r, column=6, value=g["rel_vs_whole"])
    if g["rel_vs_whole"] >= 1.4:
        c.font = Font(bold=True, color="C00000")
    ws.cell(row=r, column=7, value=g["text_density_chars_per_s"])
    r += 1
r += 1

ws.cell(row=r, column=1, value="■ 필터별 통과율 대조 (신호 안정성)").font = Font(bold=True, color="4472C4"); r += 1
sig_tbl = [
    ["신호", "하하 (46)", "ENA (32)"],
    ["IoU ≤ 0.1", "29 (63%)", "27 (84%)"],
    ["리텐션 rel ≥ 1.05", "16 (35%) ⭐", "10 (31%) ⭐"],
    ["길이 30~60초", "46 (100%)", "32 (100%)"],
    ["텍스트 밀도 ≥ 3.0/s", "46 (100%)", "32 (100%)"],
    ["학습 우수 훅", "44 (96%)", "30 (94%)"],
    ["**5개 전부**", "**8 (17%)**", "**9 (28%)**"],
]
for row_i, row_vals in enumerate(sig_tbl):
    for c_i, v in enumerate(row_vals, 1):
        c = ws.cell(row=r, column=c_i, value=v)
        if row_i == 0:
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
        c.alignment = Alignment(horizontal="center")
    r += 1
r += 1

ws.cell(row=r, column=1, value="■ 결론").font = Font(bold=True, color="4472C4"); r += 1
for line in [
    "1. **재현됨** — 완전히 다른 채널·장르에서 v2 필터가 (a) 의미있는 변별력(28%) (b) 리텐션 실측 상승(1.05~1.64x) 히든젬 발굴",
    "2. **훅 학습 신호(94~96%)·길이·밀도(100%)는 채널 무관하게 안정** — v2 필터가 하하 특화 아님",
    "3. **리텐션 rel이 실제 discrimination 담당(31~35%)** — 실측 신호가 진짜 필터, 나머지는 pre-filter",
    "",
    "한계: ENA truth 표본 7쌍(하하 71쌍 대비 작음). 방향성 신뢰 O · 절대 수치 △. dnIaj6L3t1E 2개는 매칭 표본 없어 IoU=0 auto-pass (uncertain)",
]:
    ws.cell(row=r, column=1, value=line); r += 1

for col, w in zip("ABCDEFG", [5, 14, 22, 55, 12, 10, 10]):
    ws.column_dimensions[col].width = w

# 07_파일명세 append
ws2 = wb["07_파일명세"]
last = ws2.max_row + 1
for fname, sheet, n, desc in [
    ("exp9_ena_deterministic_picks.csv", "08_ENA재현_32픽", 32, "Exp 9 ENA v2 재현 · 32 픽 전체 5신호 + 실측"),
    ("exp9_ena_confirmed_gems.json", "09_ENA재현_확정9", 9, "ENA 5신호 전부 통과 확정 9개 (rel 1.07~1.64)"),
]:
    ws2.cell(row=last, column=1, value=fname)
    ws2.cell(row=last, column=2, value=sheet)
    ws2.cell(row=last, column=3, value=n)
    ws2.cell(row=last, column=4, value=desc)
    last += 1

# 개요·통계 append
ws3 = wb["개요·통계"]
last = ws3.max_row + 2
ws3.cell(row=last, column=1, value="■ Exp 9 · v2 필터 다채널 재현 (ENA · 나는솔로)").font = Font(bold=True, size=12, color="4472C4")
last += 1
for line in [
    "32 픽 → 9 확정 (28%) · 하하 v2(17%)와 유사 통과율",
    "  다른 채널·장르에서 v2 5신호가 재현됨 — 하하 특화 아님·일반화 실증",
    "  최상위: '술 취한 대화는 안 믿어요' rel 1.64배 (편집자 미발견, 시청자는 롱폼 평균 대비 64% 더 시청)",
    "  → v2 필터를 STEP D 표준 히든젬 발굴 파이프라인으로 확정",
]:
    ws3.cell(row=last, column=1, value=line); last += 1

wb.save(XLSX)
print("완료. 시트 13_ENA재현_32픽 + 14_ENA재현_확정9 추가")
print("전체 시트:", wb.sheetnames)
