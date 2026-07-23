"""통합 xlsx에 Exp 9 드나드나 시트 추가 + 3채널 대조 요약."""
import csv
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8")
REPO = Path(r"C:\Users\STEPAI05\STEPD-repo")
XLSX = REPO / "바우처_결과보고_2026" / "제출서류_완성" / "[스텝에이아이] 결과물 증빙 데이터셋_통합.xlsx"
DATA = REPO / "바우처_결과보고_2026" / "증빙_데이터셋" / "실험자료"

wb = load_workbook(XLSX)
for name in ("10_드나드나재현_31픽", "11_드나드나재현_확정9", "12_3채널대조_v2재현"):
    if name in wb.sheetnames:
        del wb[name]

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="4472C4")
pass_fill = PatternFill("solid", fgColor="C6EFCE")

# --- 10_드나드나재현_31픽 ---
ws = wb.create_sheet("10_드나드나재현_31픽")
rows = list(csv.DictReader(open(DATA / "exp9_dna_deterministic_picks.csv", encoding="utf-8-sig")))
if rows:
    fields = list(rows[0].keys())
    for i, h in enumerate(fields, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, k in enumerate(fields, 1):
            v = row[k]
            if k.startswith("pass_") or k == "all_pass":
                v = (v == "True")
                c = ws.cell(row=r_idx, column=c_idx, value="✅" if v else "❌")
                c.alignment = Alignment(horizontal="center")
                if v: c.fill = pass_fill
            else:
                try:
                    v = float(v) if "." in str(v) or k in ("appeal",) else int(v) if str(v).lstrip("-").isdigit() else v
                except Exception:
                    pass
                ws.cell(row=r_idx, column=c_idx, value=v)
        if row.get("all_pass") == "True":
            for c_idx in range(1, len(fields)+1):
                if not ws.cell(row=r_idx, column=c_idx).fill.fgColor.rgb or ws.cell(row=r_idx, column=c_idx).fill.fgColor.rgb == "00000000":
                    ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJKLMNOPQR", [14, 10, 10, 8, 45, 10, 8, 12, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- 11_드나드나재현_확정9 ---
ws = wb.create_sheet("11_드나드나재현_확정9")
r = 1
ws.cell(row=r, column=1, value="[Exp 9 · 드나드나] 4신호 통과 히든젬 (드라마 · #허수아비 시리즈)").font = Font(bold=True, size=13); r += 1
ws.cell(row=r, column=1, value="장르: 드라마 — 3장르 재현(꽐라예능·연애리얼리티·드라마)의 세 번째. 매칭 truth 부재로 IoU auto-pass").font = Font(italic=True, color="595959"); r += 2

ws.cell(row=r, column=1, value="■ 확정 9개 (rel 순)").font = Font(bold=True, color="4472C4"); r += 1
headers = ["#", "롱폼", "구간", "제목", "훅", "rel", "밀도"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=r, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
r += 1
gems = json.load(open(DATA / "exp9_dna_confirmed_gems.json", encoding="utf-8"))
for i, g in enumerate(gems, 1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=g["long"])
    ws.cell(row=r, column=3, value=f"{g['start']:.0f}s~{g['end']:.0f}s ({g['seg_len']:.0f}초)")
    ws.cell(row=r, column=4, value=g["title"])
    ws.cell(row=r, column=5, value=g["hook"])
    c = ws.cell(row=r, column=6, value=g["rel_vs_whole"])
    if g["rel_vs_whole"] >= 1.4:
        c.font = Font(bold=True, color="C00000")
    ws.cell(row=r, column=7, value=g["text_density_chars_per_s"])
    r += 1
for col, w in zip("ABCDEFG", [5, 14, 22, 55, 12, 10, 10]):
    ws.column_dimensions[col].width = w

# --- 12_3채널대조_v2재현 ---
ws = wb.create_sheet("12_3채널대조_v2재현")
r = 1
ws.cell(row=r, column=1, value="[Exp 9] 3채널 재현성 대조 — v2 5신호 필터 다장르 일반화").font = Font(bold=True, size=14); r += 1
ws.cell(row=r, column=1, value="꽐라예능(하하) · 연애리얼리티(ENA) · 드라마(드나드나) — v2 통과율이 유사 범위(17~29%)로 안정 · 리텐션이 진짜 discriminator").font = Font(italic=True, color="595959"); r += 2

# 표1: 통과율
ws.cell(row=r, column=1, value="■ 3채널 통과율 대조").font = Font(bold=True, color="4472C4"); r += 1
tbl = [
    ["채널", "장르", "픽", "통과", "통과율", "최상위 rel"],
    ["하하 (Exp 8 v2)", "꽐라예능", 46, 8, "17%", 1.75],
    ["ENA (Exp 9)", "연애리얼리티", 32, 9, "28%", 1.64],
    ["드나드나 (Exp 9)", "드라마 (#허수아비)", 31, 9, "29%", 1.99],
]
for row_i, row_vals in enumerate(tbl):
    for c_i, v in enumerate(row_vals, 1):
        c = ws.cell(row=r, column=c_i, value=v)
        if row_i == 0:
            c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    r += 1
r += 1

# 표2: 신호별
ws.cell(row=r, column=1, value="■ 필터별 통과율 (신호 안정성)").font = Font(bold=True, color="4472C4"); r += 1
sig_tbl = [
    ["신호", "하하 46", "ENA 32", "드나드나 31", "해석"],
    ["IoU ≤ 0.1", "63%", "84%", "100%*", "*드나드나 truth 없어 auto-pass"],
    ["리텐션 rel ≥ 1.05 ⭐", "35%", "31%", "29%", "🎯 진짜 discriminator (29~35%)"],
    ["길이 30~60초", "100%", "100%", "96%", "pre-filter, 채널 무관 안정"],
    ["밀도 ≥ 3.0/s", "100%", "100%", "93%", "pre-filter, 채널 무관 안정"],
    ["학습 우수 훅", "96%", "94%", "93%", "pre-filter, 채널 무관 안정"],
    ["5개 전부", "17%", "28%", "29%", "**최종 통과 유사 범위**"],
]
for row_i, row_vals in enumerate(sig_tbl):
    for c_i, v in enumerate(row_vals, 1):
        c = ws.cell(row=r, column=c_i, value=v)
        if row_i == 0:
            c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    r += 1
r += 1

# 결론
ws.cell(row=r, column=1, value="■ 결론").font = Font(bold=True, color="4472C4"); r += 1
for line in [
    "1. **3장르 재현 확정** — v2 필터 통과율이 17%·28%·29%로 유사 범위. 방법론이 채널·장르 종속 아님",
    "2. **리텐션 rel이 진짜 discriminator** — 3채널 모두 29~35% 안정적 변별력. 나머지는 pre-filter(93~100%)",
    "3. **최상위 히든젬 rel 1.99배** (드나드나 '내 동생 살려내') · 3채널 통틀어 최고치 · 시청자가 롱폼 평균 2배 오래 봄",
    "",
    "한계: 드나드나 매칭 표본 0 → IoU 신호 무효(4신호 판정). 상위 shorts·상위 longs가 다른 롱폼 계열",
    "다음: 프로덕션 자동 매칭(match.align 배치)로 규모 확보 → truth 대량화 · Hit@N 정식 측정",
]:
    ws.cell(row=r, column=1, value=line); r += 1

for col, w in zip("ABCDEF", [18, 16, 12, 12, 15, 40]):
    ws.column_dimensions[col].width = w

# 07_파일명세 append
ws2 = wb["07_파일명세"]
last = ws2.max_row + 1
for fname, sheet, n, desc in [
    ("exp9_dna_deterministic_picks.csv", "10_드나드나재현_31픽", 31, "Exp 9 드나드나 v2 재현 · 31 픽 (매칭 truth 없어 IoU auto-pass)"),
    ("exp9_dna_confirmed_gems.json", "11_드나드나재현_확정9", 9, "드나드나 4신호 전부 통과 확정 9개 (rel 1.05~1.99)"),
]:
    ws2.cell(row=last, column=1, value=fname)
    ws2.cell(row=last, column=2, value=sheet)
    ws2.cell(row=last, column=3, value=n)
    ws2.cell(row=last, column=4, value=desc)
    last += 1

# 개요·통계 append
ws3 = wb["개요·통계"]
last = ws3.max_row + 2
ws3.cell(row=last, column=1, value="■ Exp 9 · 3채널 재현성 확정 (하하·ENA·드나드나)").font = Font(bold=True, size=12, color="4472C4")
last += 1
for line in [
    "3장르 대조: 하하 17% · ENA 28% · 드나드나 29% — v2 통과율 유사 범위, 채널·장르 종속 X",
    "  리텐션 rel이 진짜 discriminator(29~35%) · 나머지 신호는 채널 무관 안정(93~100%)",
    "  최상위: 드나드나 '내 동생 살려내' rel 1.99배 (롱폼 평균 2배, 3채널 최고치)",
    "  → v2 필터를 STEP D 표준 히든젬 발굴 파이프라인으로 확정",
]:
    ws3.cell(row=last, column=1, value=line); last += 1

wb.save(XLSX)
print("완료. 시트 추가: 10_드나드나재현_31픽 · 11_드나드나재현_확정9 · 12_3채널대조_v2재현")
print("전체 시트:", wb.sheetnames)
